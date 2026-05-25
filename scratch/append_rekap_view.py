import os

views_file = "/Users/muhakbaryasin/Desktop/master/kpukonut/simpadi/perjalanan/views.py"

content_to_append = """

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def rekap_perjadin_view(request):
    from master_data.models import Pegawai
    from .models import PerjalananDinas
    import openpyxl
    import os
    from django.conf import settings
    
    pegawai_id = request.GET.get('pegawai_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    export = request.GET.get('export')
    
    pegawais = Pegawai.objects.all().order_by('nama')
    qs = PerjalananDinas.objects.filter(status=PerjalananDinas.Status.APPROVED)
    
    if pegawai_id:
        qs = qs.filter(pegawai_id=pegawai_id)
    if start_date:
        qs = qs.filter(surat_tugas__tanggal_berangkat__gte=start_date)
    if end_date:
        qs = qs.filter(surat_tugas__tanggal_kembali__lte=end_date)
        
    qs = qs.select_related('surat_tugas', 'pegawai', 'biaya').order_by('surat_tugas__tanggal_berangkat')
    
    data_rekap = []
    for idx, p in enumerate(qs, start=1):
        st = p.surat_tugas
        pegawai = p.pegawai
        biaya = p.biaya if hasattr(p, 'biaya') else None
        
        uang_harian_riil = biaya.uang_harian_riil if biaya else 0
        uang_transport_riil = biaya.uang_transport_riil if biaya else 0
        uang_penginapan_riil = biaya.uang_penginapan_riil if biaya else 0
        uang_representasi_riil = biaya.uang_representasi_riil if biaya else 0
        uang_tiket_riil = biaya.uang_tiket_riil if biaya else 0
        
        total_biaya_lainnya = uang_transport_riil + uang_representasi_riil
        total_biaya = uang_harian_riil + total_biaya_lainnya + uang_penginapan_riil + uang_tiket_riil
        
        item = {
            'no': idx,
            'opd': 'KPU PROVINSI SULAWESI TENGGARA',
            'jenis_perjadin': st.get_jenis_perjalanan_display() if st else '',
            'no_bukti': '',
            'tgl_bukti': '',
            'nama_pegawai': pegawai.nama if pegawai else '',
            'nip': pegawai.nip if pegawai else '',
            'no_st': st.nomor_surat if st else '',
            'tgl_st': st.tgl_surat if st else '',
            'keperluan': st.perihal if st else '',
            'jabatan': pegawai.jabatan if pegawai else '',
            'tujuan': st.tempat_tujuan if st else '',
            'visum': '',
            'tgl_berangkat': st.tanggal_berangkat if st else '',
            'tgl_kembali': st.tanggal_kembali if st else '',
            'lama_hari': p.durasi_hari if p else '',
            'uang_harian_riil': uang_harian_riil,
            'uang_transport_riil': uang_transport_riil,
            'uang_representasi_riil': uang_representasi_riil,
            'total_biaya_lainnya': total_biaya_lainnya,
            'uang_penginapan_riil': uang_penginapan_riil,
            'uang_tiket_riil': uang_tiket_riil,
            'total_biaya': total_biaya,
        }
        data_rekap.append(item)
        
    if export == 'excel':
        template_path = os.path.join(settings.MEDIA_ROOT, 'Pokok Perjadin.xlsx')
        if not os.path.exists(template_path):
            return HttpResponse("Template Excel Pokok Perjadin.xlsx tidak ditemukan di folder media.")
            
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        row_idx = 8
        for item in data_rekap:
            ws.cell(row=row_idx, column=1, value=item['no'])
            ws.cell(row=row_idx, column=2, value=item['opd'])
            ws.cell(row=row_idx, column=3, value=item['jenis_perjadin'])
            ws.cell(row=row_idx, column=4, value=item['no_bukti'])
            ws.cell(row=row_idx, column=5, value=item['tgl_bukti'])
            ws.cell(row=row_idx, column=6, value=item['nama_pegawai'])
            ws.cell(row=row_idx, column=7, value=item['nip'])
            ws.cell(row=row_idx, column=8, value=item['no_st'])
            ws.cell(row=row_idx, column=9, value=item['tgl_st'].strftime("%d/%m/%Y") if item['tgl_st'] else "")
            ws.cell(row=row_idx, column=10, value=item['keperluan'])
            ws.cell(row=row_idx, column=11, value=item['jabatan'])
            ws.cell(row=row_idx, column=12, value=item['tujuan'])
            ws.cell(row=row_idx, column=13, value=item['visum'])
            ws.cell(row=row_idx, column=14, value=item['tgl_berangkat'].strftime("%d/%m/%Y") if item['tgl_berangkat'] else "")
            ws.cell(row=row_idx, column=15, value=item['tgl_kembali'].strftime("%d/%m/%Y") if item['tgl_kembali'] else "")
            ws.cell(row=row_idx, column=16, value=item['lama_hari'])
            ws.cell(row=row_idx, column=18, value=item['uang_harian_riil'])
            ws.cell(row=row_idx, column=19, value=item['uang_transport_riil'])
            ws.cell(row=row_idx, column=20, value=item['uang_representasi_riil'])
            ws.cell(row=row_idx, column=22, value=item['total_biaya_lainnya'])
            ws.cell(row=row_idx, column=28, value=item['uang_penginapan_riil'])
            ws.cell(row=row_idx, column=43, value=item['uang_tiket_riil'])
            ws.cell(row=row_idx, column=44, value=item['total_biaya'])
            row_idx += 1
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rekap_Perjadin.xlsx"'
        wb.save(response)
        return response

    context = {
        'pegawais': pegawais,
        'data_rekap': data_rekap,
        'selected_pegawai': int(pegawai_id) if pegawai_id else '',
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'perjalanan/admin_rekap_perjadin.html', context)
"""

with open(views_file, "a") as f:
    f.write(content_to_append)
