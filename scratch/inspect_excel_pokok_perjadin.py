import openpyxl
import os

file_path = "/Users/muhakbaryasin/Desktop/master/kpukonut/simpadi/media/Pokok Perjadin.xlsx"

if not os.path.exists(file_path):
    print("File not found")
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            print(row)
