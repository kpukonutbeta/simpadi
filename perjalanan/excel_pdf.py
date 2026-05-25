"""Render an Excel workbook to PDF while preserving its cell layout as much as possible.

This is a lightweight, dependency-free fallback for environments where
LibreOffice / soffice is not installed.
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, Optional, Tuple
from xml.sax.saxutils import escape

from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph

try:
    from openpyxl.utils.units import column_width_to_pixels
except Exception:  # pragma: no cover - fallback if openpyxl internals differ
    column_width_to_pixels = None


DEFAULT_FONT_SIZE = 10
DEFAULT_ROW_HEIGHT = 15.0
DEFAULT_COL_WIDTH = 8.43
PAGE_MARGIN_LEFT = 44.64   # 0.62 in
PAGE_MARGIN_RIGHT = 28.35  # 0.393701 in
PAGE_MARGIN_TOP = 53.86    # 0.748031 in
PAGE_MARGIN_BOTTOM = 53.86
PADDING = 2.0


FONT_MAP = {
    "helvetica": {
        (False, False): "Helvetica",
        (True, False): "Helvetica-Bold",
        (False, True): "Helvetica-Oblique",
        (True, True): "Helvetica-BoldOblique",
    },
    "arial": {
        (False, False): "Helvetica",
        (True, False): "Helvetica-Bold",
        (False, True): "Helvetica-Oblique",
        (True, True): "Helvetica-BoldOblique",
    },
    "calibri": {
        (False, False): "Helvetica",
        (True, False): "Helvetica-Bold",
        (False, True): "Helvetica-Oblique",
        (True, True): "Helvetica-BoldOblique",
    },
    "times new roman": {
        (False, False): "Times-Roman",
        (True, False): "Times-Bold",
        (False, True): "Times-Italic",
        (True, True): "Times-BoldItalic",
    },
    "courier new": {
        (False, False): "Courier",
        (True, False): "Courier-Bold",
        (False, True): "Courier-Oblique",
        (True, True): "Courier-BoldOblique",
    },
}


BORDER_WIDTHS = {
    None: 0,
    "hair": 0.25,
    "thin": 0.5,
    "medium": 1.0,
    "thick": 1.5,
    "dashed": 0.5,
    "dotted": 0.5,
    "dashDot": 0.5,
    "dashDotDot": 0.5,
    "double": 1.0,
}


def _column_width_to_points(width: Optional[float]) -> float:
    width = DEFAULT_COL_WIDTH if width is None else float(width)
    if column_width_to_pixels is not None:
        pixels = column_width_to_pixels(width)
    else:
        pixels = int(width * 7 + 5)
    return pixels * 72.0 / 96.0


def _row_height_to_points(height: Optional[float]) -> float:
    return float(height) if height is not None else DEFAULT_ROW_HEIGHT


def _color_from_rgb(rgb: Optional[str]):
    if not rgb:
        return None
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    try:
        return colors.HexColor(f"#{rgb}")
    except Exception:
        return None


def _font_name(font) -> str:
    family = (getattr(font, "name", None) or "Helvetica").lower()
    bold = bool(getattr(font, "bold", False))
    italic = bool(getattr(font, "italic", False))
    for key, mapping in FONT_MAP.items():
        if key in family:
            return mapping[(bold, italic)]
    return FONT_MAP["helvetica"][(bold, italic)]


def _border_spec(side) -> Tuple[float, Optional[Tuple[int, int]]]:
    style = getattr(side, "style", None)
    width = BORDER_WIDTHS.get(style, 0.5 if style else 0)
    dash = None
    if style in {"dashed", "dashDot", "dashDotDot"}:
        dash = (3, 2)
    elif style == "dotted":
        dash = (1, 2)
    return width, dash


def _merged_map(ws):
    merged = {}
    skip = set()
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        merged[(min_row, min_col)] = (min_row, min_col, max_row, max_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    skip.add((r, c))
    return merged, skip


def _sheet_dimensions(ws):
    max_col = ws.max_column
    max_row = ws.max_row
    col_widths = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        col_widths.append(_column_width_to_points(dim.width))
    row_heights = []
    for row_idx in range(1, max_row + 1):
        dim = ws.row_dimensions[row_idx]
        row_heights.append(_row_height_to_points(dim.height))
    return col_widths, row_heights


def _box_positions(sizes: Iterable[float]):
    pos = [0.0]
    total = 0.0
    for value in sizes:
        total += value
        pos.append(total)
    return pos


def _wrap_text_to_fit(text: str, font_name: str, font_size: float, width: float):
    if not text:
        return ""
    lines = []
    for paragraph in text.splitlines() or [text]:
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            if stringWidth(candidate, font_name, font_size) <= width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return "<br/>".join(escape(line) for line in lines)


def _draw_paragraph(c, text, x, y, width, height, font_name, font_size, color, alignment, valign):
    if text is None:
        return
    raw = str(text)
    if raw == "":
        return

    if alignment == "center":
        ta = TA_CENTER
    elif alignment == "right":
        ta = TA_RIGHT
    else:
        ta = TA_LEFT

    inner_w = max(width - (PADDING * 2), 1)
    inner_h = max(height - (PADDING * 2), 1)

    chosen_size = float(font_size or DEFAULT_FONT_SIZE)
    for _ in range(5):
        style = ParagraphStyle(
            "cell-style",
            fontName=font_name,
            fontSize=chosen_size,
            leading=max(chosen_size * 1.15, chosen_size + 0.5),
            textColor=color or colors.black,
            alignment=ta,
            wordWrap="CJK",
        )
        para = Paragraph(_wrap_text_to_fit(raw, font_name, chosen_size, inner_w), style)
        pw, ph = para.wrap(inner_w, inner_h)
        if ph <= inner_h or chosen_size <= 6:
            break
        chosen_size -= 1

    pw, ph = para.wrap(inner_w, inner_h)
    if valign == "middle":
        py = y + (height - ph) / 2.0
    elif valign == "top":
        py = y + height - ph - PADDING
    else:
        py = y + PADDING

    px = x + PADDING
    if alignment == "center":
        px = x + max((width - pw) / 2.0, PADDING)
    elif alignment == "right":
        px = x + max(width - pw - PADDING, PADDING)

    para.drawOn(c, px, py)


def _draw_cell_border(c, x, y, width, height, border):
    if border is None:
        return

    left_w, left_dash = _border_spec(border.left)
    right_w, right_dash = _border_spec(border.right)
    top_w, top_dash = _border_spec(border.top)
    bottom_w, bottom_dash = _border_spec(border.bottom)

    def _line(x1, y1, x2, y2, line_width, dash):
        if line_width <= 0:
            return
        c.saveState()
        c.setLineWidth(line_width)
        if dash:
            c.setDash(*dash)
        c.line(x1, y1, x2, y2)
        c.restoreState()

    _line(x, y, x, y + height, left_w, left_dash)
    _line(x + width, y, x + width, y + height, right_w, right_dash)
    _line(x, y + height, x + width, y + height, top_w, top_dash)
    _line(x, y, x + width, y, bottom_w, bottom_dash)


def _draw_fill(c, x, y, width, height, fill):
    if fill is None:
        return
    if getattr(fill, "patternType", None) != "solid":
        return
    fg = getattr(fill, "fgColor", None)
    color = None
    if fg is not None:
        if getattr(fg, "type", None) == "rgb":
            color = _color_from_rgb(fg.rgb)
        elif getattr(fg, "index", None) is not None:
            color = None
    if color is None:
        return
    c.saveState()
    c.setFillColor(color)
    c.setStrokeColor(color)
    c.rect(x, y, width, height, stroke=0, fill=1)
    c.restoreState()


def _image_bounds(ws, img, left_margin, top_margin, scale, col_offsets, row_offsets, col_widths, row_heights, sheet_top):
    anchor = getattr(img, "anchor", None)
    if anchor is None or not hasattr(anchor, "_from") or not hasattr(anchor, "to"):
        return None

    frm = anchor._from
    to = anchor.to

    def emu_to_points(emu):
        return (emu or 0) / 12700.0

    x1 = left_margin + (col_offsets[frm.col] + emu_to_points(getattr(frm, "colOff", 0))) * scale
    y1_top = sheet_top - (row_offsets[frm.row] + emu_to_points(getattr(frm, "rowOff", 0))) * scale
    x2 = left_margin + (col_offsets[to.col] + emu_to_points(getattr(to, "colOff", 0))) * scale
    y2_bottom = sheet_top - (row_offsets[to.row] + emu_to_points(getattr(to, "rowOff", 0))) * scale

    width = max(x2 - x1, 1)
    height = max(y1_top - y2_bottom, 1)
    return x1, y2_bottom, width, height


def render_worksheet_to_pdf_bytes(ws) -> bytes:
    """Render a worksheet to PDF bytes.

    The function is intentionally pragmatic: it preserves the template layout
    closely enough for administrative documents, including solid fills,
    borders, merged cells, text alignment, and embedded images.
    """

    page_width, page_height = A4
    col_widths, row_heights = _sheet_dimensions(ws)
    merged, skip = _merged_map(ws)
    col_offsets = _box_positions(col_widths)
    row_offsets = _box_positions(row_heights)

    sheet_width = sum(col_widths)
    sheet_height = sum(row_heights)

    avail_width = page_width - PAGE_MARGIN_LEFT - PAGE_MARGIN_RIGHT
    avail_height = page_height - PAGE_MARGIN_TOP - PAGE_MARGIN_BOTTOM
    scale = min(avail_width / sheet_width if sheet_width else 1.0, avail_height / sheet_height if sheet_height else 1.0)
    sheet_top = page_height - PAGE_MARGIN_TOP

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4, pageCompression=1)
    c.setTitle(getattr(ws, "title", "Worksheet"))

    # White base page.
    c.setFillColor(colors.white)
    c.rect(0, 0, page_width, page_height, stroke=0, fill=1)

    # Draw cells from back to front.
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            if (row_idx, col_idx) in skip:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None and (row_idx, col_idx) not in merged:
                continue

            if (row_idx, col_idx) in merged:
                min_row, min_col, max_row, max_col = merged[(row_idx, col_idx)]
                x = PAGE_MARGIN_LEFT + col_offsets[min_col - 1] * scale
                y_top = sheet_top - row_offsets[min_row - 1] * scale
                width = (col_offsets[max_col] - col_offsets[min_col - 1]) * scale
                height = (row_offsets[max_row] - row_offsets[min_row - 1]) * scale
            else:
                x = PAGE_MARGIN_LEFT + col_offsets[col_idx - 1] * scale
                y_top = sheet_top - row_offsets[row_idx - 1] * scale
                width = col_widths[col_idx - 1] * scale
                height = row_heights[row_idx - 1] * scale

            y = y_top - height

            _draw_fill(c, x, y, width, height, cell.fill)
            _draw_cell_border(c, x, y, width, height, cell.border)

            if cell.value is not None:
                font_name = _font_name(cell.font)
                font_size = float(cell.font.sz or DEFAULT_FONT_SIZE)
                color = _color_from_rgb(cell.font.color.rgb) if getattr(cell.font, "color", None) and getattr(cell.font.color, "type", None) == "rgb" else None
                alignment = getattr(cell.alignment, "horizontal", None) or "left"
                valign = getattr(cell.alignment, "vertical", None) or "middle"
                _draw_paragraph(c, cell.value, x, y, width, height, font_name, font_size, color, alignment, valign)

    # Draw embedded images last so they appear above cell backgrounds.
    for img in getattr(ws, "_images", []):
        try:
            bounds = _image_bounds(ws, img, PAGE_MARGIN_LEFT, PAGE_MARGIN_TOP, scale, col_offsets, row_offsets, col_widths, row_heights, sheet_top)
            if not bounds:
                continue
            x, y, width, height = bounds
            data = img._data()
            if not data:
                continue
            reader = ImageReader(BytesIO(data))
            c.drawImage(reader, x, y, width=width, height=height, preserveAspectRatio=True, mask='auto')
        except Exception:
            # Ignore broken images; the data cells remain valid.
            continue

    c.showPage()
    c.save()
    return buf.getvalue()


def render_workbook_to_pdf_bytes(workbook) -> bytes:
    """Render the workbook's active worksheet to PDF bytes."""

    ws = workbook.active
    return render_worksheet_to_pdf_bytes(ws)

