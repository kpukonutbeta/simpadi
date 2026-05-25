from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.contrib.admin.views.decorators import staff_member_required
from .models import SuratTugas, PerjalananDinas
from .forms import PerjalananDinasForm, BiayaPerjalananFormSet, BerkasPerjalananFormSet, HarianPerjalananFormSet
from master_data.models import Provinsi, Anggaran

def get_pegawai_by_surat_tugas(request):
    surat_tugas_id = request.GET.get('surat_tugas_id')
    if not surat_tugas_id:
        return JsonResponse({'pegawai': []})
    
    try:
        surat = SuratTugas.objects.get(id=surat_tugas_id)
        pegawai_list = [
            {'id': p.id, 'nama': f"{p.nama} ({p.nip})"}
            for p in surat.pegawai.all()
        ]
        return JsonResponse({'pegawai': pegawai_list})
    except (SuratTugas.DoesNotExist, ValueError):
        return JsonResponse({'pegawai': []})

@login_required
def ajukan_perjadin(request, surat_tugas_id):
    # Ensure user has a Pegawai profile
    if not hasattr(request.user, 'pegawai_profile'):
        messages.error(request, "Akun Anda belum terhubung dengan data Pegawai.")
        return redirect('core:dashboard')
    
    pegawai = request.user.pegawai_profile
    surat_tugas = get_object_or_404(SuratTugas, id=surat_tugas_id)
    
    # Ensure Pegawai is assigned to this Surat Tugas
    if not surat_tugas.pegawai.filter(id=pegawai.id).exists():
        messages.error(request, "Anda tidak ditugaskan dalam Surat Tugas ini.")
        return redirect('core:dashboard')

    # Check if PerjalananDinas already exists for this ST and Pegawai
    # This enables syncing between Admin-created drafts and Pegawai
    perjadin_instance, created = PerjalananDinas.objects.get_or_create(
        surat_tugas=surat_tugas, 
        pegawai=pegawai,
        defaults={'status': PerjalananDinas.Status.DRAFT}
    )

    # Allow editing for DRAFT and REJECTED status only
    is_editable = perjadin_instance.status in [PerjalananDinas.Status.DRAFT, PerjalananDinas.Status.REJECTED]
    
    # If status is locked (not DRAFT/REJECTED), show view-only mode
    if not is_editable and request.method == 'POST':
        messages.warning(request, f"SPD ini sudah dalam status {perjadin_instance.get_status_display()} dan tidak dapat diubah lagi.")
        return redirect('core:dashboard')

    if request.method == 'POST':
        form = PerjalananDinasForm(request.POST, instance=perjadin_instance, user=request.user)
        
        # SANGAT PENTING: Tempelkan instance sebelum validasi agar clean() di model tidak error
        form.instance.surat_tugas = surat_tugas
        form.instance.pegawai = pegawai
        
        biaya_formset = BiayaPerjalananFormSet(request.POST, instance=perjadin_instance)
        berkas_formset = BerkasPerjalananFormSet(request.POST, request.FILES, instance=perjadin_instance)
        harian_formset = HarianPerjalananFormSet(request.POST, instance=perjadin_instance)
        
        if not request.user.is_staff:
            for h_form in harian_formset:
                h_form.fields['provinsi'].disabled = True
                h_form.fields['jenis_harian'].disabled = True

        if form.is_valid() and biaya_formset.is_valid() and berkas_formset.is_valid() and harian_formset.is_valid():
            try:
                with transaction.atomic():
                    # Save main form
                    perjadin = form.save(commit=False)
                    perjadin.surat_tugas = surat_tugas
                    perjadin.pegawai = pegawai
                    
                    # Check which action button was clicked
                    action = request.POST.get('action', 'save')
                    
                    # If submit_verification action, change status to PENDING
                    if action == 'submit_verification':
                        perjadin.status = PerjalananDinas.Status.PENDING
                        message_type = "SPD berhasil diajukan untuk verifikasi"
                    else:
                        # Keep status as DRAFT if it was new, or keep existing status if it was REJECTED
                        if not perjadin.status:
                            perjadin.status = PerjalananDinas.Status.DRAFT
                        message_type = "Data Perjalanan Dinas berhasil disimpan"
                    
                    perjadin.save()
                    
                    # Save formsets: harian_formset first so that when BiayaPerjalanan calculates, it reads the updated daily transit SBM
                    harian_formset.instance = perjadin
                    harian_formset.save()

                    # Save formsets: berkas_formset next so that BiayaPerjalanan can read the saved berkas nominals
                    berkas_formset.instance = perjadin
                    berkas_formset.save()
                    
                    biaya_formset.instance = perjadin
                    biaya_formset.save()
                    
                messages.success(request, message_type + ".")
                if action == 'save':
                    return redirect('perjalanan:ajukan_perjadin', surat_tugas_id=surat_tugas.id)
                else:
                    return redirect('core:dashboard')
            except Exception as e:
                messages.error(request, f"Terjadi kesalahan saat menyimpan: {e}")
        else:
            messages.error(request, "Terdapat kesalahan pada isian form. Silakan periksa kembali.")
    else:
        form = PerjalananDinasForm(instance=perjadin_instance, user=request.user)
        biaya_formset = BiayaPerjalananFormSet(instance=perjadin_instance)
        berkas_formset = BerkasPerjalananFormSet(instance=perjadin_instance)
        harian_formset = HarianPerjalananFormSet(instance=perjadin_instance)
        
        if not request.user.is_staff:
            for h_form in harian_formset:
                h_form.fields['provinsi'].disabled = True
                h_form.fields['jenis_harian'].disabled = True
        
        # Disable all form fields if status is not editable (not DRAFT or REJECTED)
        if not is_editable:
            for field in form.fields.values():
                field.disabled = True
            for b_form in biaya_formset:
                for field in b_form.fields.values():
                    field.disabled = True
            for bk_form in berkas_formset:
                for field in bk_form.fields.values():
                    field.disabled = True
            for h_form in harian_formset:
                for field in h_form.fields.values():
                    field.disabled = True

    from master_data.models import JenisBerkas
    jenis_berkas_nominal = list(JenisBerkas.objects.filter(nominal_biaya=True).values_list('id', flat=True))
    jenis_berkas_wajib = list(JenisBerkas.objects.filter(wajib=True).values_list('id', flat=True))
    
    from django.db.models import Q
    jenis_berkas_penginapan = list(JenisBerkas.objects.filter(
        kategori_biaya__in=['penginapan', 'penginapan_30', 'penginapan_fb_luar', 'penginapan_fb_dalam']
    ).values_list('id', flat=True))
    
    jenis_berkas_fb_luar = list(JenisBerkas.objects.filter(
        kategori_biaya='penginapan_fb_luar'
    ).values_list('id', flat=True))
    
    jenis_berkas_fb_dalam = list(JenisBerkas.objects.filter(
        kategori_biaya='penginapan_fb_dalam'
    ).values_list('id', flat=True))

    jenis_berkas_tiket_pesawat = list(JenisBerkas.objects.filter(
        kategori_biaya='transportasi_pesawat'
    ).values_list('id', flat=True))

    jenis_berkas_transportasi = list(JenisBerkas.objects.filter(
        kategori_biaya__in=['transportasi', 'transportasi_pesawat']
    ).values_list('id', flat=True))

    breakdown_initial = None
    if perjadin_instance and hasattr(perjadin_instance, 'biaya') and perjadin_instance.biaya:
        try:
            breakdown_initial = perjadin_instance.biaya.calculate_breakdown()
        except Exception:
            pass

    from master_data.models import DokumenSBM
    sbm_dokumen = DokumenSBM.objects.filter(tahun=surat_tugas.tahun_sbm).first()

    context = {
        'surat_tugas': surat_tugas,
        'perjadin': perjadin_instance,
        'form': form,
        'biaya_formset': biaya_formset,
        'berkas_formset': berkas_formset,
        'harian_formset': harian_formset,
        'is_edit': True,
        'is_editable': is_editable,
        'jenis_berkas_nominal': jenis_berkas_nominal,
        'jenis_berkas_wajib': jenis_berkas_wajib,
        'jenis_berkas_penginapan': jenis_berkas_penginapan,
        'jenis_berkas_fb_luar': jenis_berkas_fb_luar,
        'jenis_berkas_fb_dalam': jenis_berkas_fb_dalam,
        'jenis_berkas_tiket_pesawat': jenis_berkas_tiket_pesawat,
        'jenis_berkas_transportasi': jenis_berkas_transportasi,
        'breakdown_initial': breakdown_initial,
        'sbm_dokumen': sbm_dokumen
    }
    return render(request, 'perjalanan/ajukan_form.html', context)

@staff_member_required
def generate_spd_bulk(request):
    selected_ids = request.session.get('selected_st_ids', [])
    if not selected_ids:
        messages.warning(request, "Tidak ada Surat Tugas yang dipilih.")
        return redirect('/admin/perjalanan/surattugas/')
    
    surat_tugas_list = SuratTugas.objects.filter(id__in=selected_ids)
    
    # Get default suffix from new config model
    from .models import PengaturanNomorSPD
    config, _ = PengaturanNomorSPD.objects.get_or_create(id=1)
    default_suffix = config.suffix_format

    if request.method == 'POST':
        count_created = 0
        count_skipped = 0
        try:
            for st in surat_tugas_list:
                for pgw in st.pegawai.all():
                    # Check if SPD already exists for this pair
                    if not PerjalananDinas.objects.filter(surat_tugas=st, pegawai=pgw).exists():
                        try:
                            with transaction.atomic():
                                # Note: nomor_spd will be generated automatically in .save()
                                PerjalananDinas.objects.create(
                                    surat_tugas=st,
                                    pegawai=pgw,
                                    status=PerjalananDinas.Status.DRAFT,
                                )
                                count_created += 1
                        except Exception as e:
                            count_skipped += 1
                    else:
                        count_skipped += 1
                            
            if count_created > 0:
                messages.success(request, f"Berhasil menerbitkan {count_created} SPD baru dengan nomor otomatis.")
            if count_skipped > 0:
                messages.info(request, f"{count_skipped} SPD dilewati karena sudah pernah diterbitkan sebelumnya.")
            if count_created == 0 and count_skipped == 0:
                messages.warning(request, "Tidak ada data SPD yang diproses.")
                
            del request.session['selected_st_ids']
            return redirect('/admin/perjalanan/perjalanandinas/')
        except Exception as e:
            messages.error(request, f"Gagal menerbitkan SPD: {e}")

    context = {
        'surat_tugas_list': surat_tugas_list,
        'default_suffix': default_suffix,
    }
    return render(request, 'perjalanan/admin_generate_spd.html', context)

@login_required
def riwayat_perjadin(request):
    if not hasattr(request.user, 'pegawai_profile'):
        messages.error(request, "Akun Anda belum terhubung dengan data Pegawai.")
        return redirect('core:dashboard')
    
    pegawai = request.user.pegawai_profile
    # Filter only those that are NOT in DRAFT (or include DRAFT if you want full history)
    # Usually history includes everything the user has done.
    history = PerjalananDinas.objects.filter(pegawai=pegawai).order_by('-surat_tugas__tanggal_berangkat', '-created_at')
    
    return render(request, 'perjalanan/riwayat_list.html', {'history': history})


import json
from datetime import datetime
from django.views.decorators.http import require_POST
from master_data.models import StandarBiaya, Pegawai, JenisBerkas

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    # Try common formats: YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY, YYYY/MM/DD
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

@login_required
@require_POST
def hitung_estimasi_ajax(request):
    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST

    # Retrieve parameters
    tanggal_berangkat_str = data.get('tanggal_berangkat')
    tanggal_kembali_str = data.get('tanggal_kembali')
    tujuan_provinsi_id = data.get('tujuan_provinsi')
    jenis_transportasi = data.get('jenis_transportasi')
    jenis_perjalanan = data.get('jenis_perjalanan')
    tahun_sbm = data.get('tahun_sbm')
    pegawai_id = data.get('pegawai_id')
    
    # If pegawai_id is not passed, use current user's profile
    if not pegawai_id and hasattr(request.user, 'pegawai_profile'):
        pegawai_obj = request.user.pegawai_profile
    elif pegawai_id:
        try:
            pegawai_obj = Pegawai.objects.get(id=pegawai_id)
        except Pegawai.DoesNotExist:
            pegawai_obj = None
    else:
        pegawai_obj = None

    tb = parse_date(tanggal_berangkat_str) if tanggal_berangkat_str else None
    tk = parse_date(tanggal_kembali_str) if tanggal_kembali_str else None
    prov = None
    if tujuan_provinsi_id:
        try:
            prov = Provinsi.objects.get(id=tujuan_provinsi_id)
        except Provinsi.DoesNotExist:
            pass

    from .models import PerjalananDinas, SuratTugas, BiayaPerjalanan

    # Construct mock SuratTugas and PerjalananDinas
    st = SuratTugas(
        tanggal_berangkat=tb,
        tanggal_kembali=tk,
        tujuan_provinsi=prov,
        jenis_perjalanan=jenis_perjalanan,
        jenis_transportasi=jenis_transportasi,
        tahun_sbm=int(tahun_sbm) if tahun_sbm else 2024
    )
    p = PerjalananDinas(
        surat_tugas=st,
        pegawai=pegawai_obj
    )
    b_obj = BiayaPerjalanan(
        perjalanan=p
    )

    berkas_payload = data.get('berkas', [])
    harian_payload = data.get('harian', [])
    breakdown = b_obj.calculate_breakdown(mock_berkas=berkas_payload, mock_harian=harian_payload)

    # Serialize breakdown_categories for AJAX
    serialized_categories = {}
    for key, category in breakdown['breakdown_categories'].items():
        serialized_items = []
        for item in category['items']:
            serialized_items.append({
                'perihal': item['perihal'],
                'no': item['no'],
                'keterangan': item['keterangan'],
                'harga': float(item['harga']),
                'kuantitas': item['kuantitas'],
                'total': float(item['total']),
                'file_url': item['file_url'],
                'file_name': item['file_name'],
            })
        serialized_categories[key] = {
            'title': category['title'],
            'items': serialized_items,
            'subtotal': float(category['subtotal']),
        }

    serialized_harian_breakdown = []
    for item in breakdown.get('harian_breakdown', []):
        serialized_harian_breakdown.append({
            'hari_ke': item['hari_ke'],
            'tanggal': item['tanggal'].strftime('%Y-%m-%d') if item['tanggal'] else "",
            'provinsi_id': item['provinsi_id'],
            'provinsi_nama': item['provinsi_nama'],
            'jenis_harian': item['jenis_harian'],
            'rate': float(item['rate']),
            'representasi': float(item['representasi']),
            'representasi_expected': float(item.get('representasi_expected', 0)),
            'representasi_label': item.get('representasi_label', ''),
        })

    # Return JSON with all float/int format for numeric items
    return JsonResponse({
        'uang_harian_riil': float(breakdown['uang_harian_riil']),
        'uang_representasi_riil': float(breakdown['uang_representasi_riil']),
        'biaya_penginapan_riil': float(breakdown['biaya_penginapan_riil']),
        'biaya_transportasi_riil': float(breakdown['biaya_transportasi_riil']),
        'total_dibayarkan': float(breakdown['total_dibayarkan']),
        'total_tidak_dibayarkan': float(breakdown['total_dana_pribadi']),
        'durasi_hari': breakdown['durasi_hari'],
        'over_limit': breakdown['over_limit'],
        'plafon_hotel': float(breakdown['sbm_plafon_hotel']),
        'sbm_uang_representasi_luar_kota': float(breakdown.get('sbm_uang_representasi_luar_kota', 0)),
        'sbm_uang_representasi_dalam_kota': float(breakdown.get('sbm_uang_representasi_dalam_kota', 0)),
        'harian_formula': breakdown['harian_formula'],
        'representasi_formula': breakdown['representasi_formula'],
        'penginapan_formula': breakdown['penginapan_formula'],
        'transport_formula': breakdown['transport_formula'],
        'breakdown_categories': serialized_categories,
        'harian_breakdown': serialized_harian_breakdown,
    })

@login_required
def get_standar_biaya_tiket_ajax(request):
    from master_data.models import StandarBiayaTiket, JenisBerkas
    tahun_sbm = request.GET.get('tahun_sbm')
    pegawai_id = request.GET.get('pegawai_id')
    
    tiket_qs = StandarBiayaTiket.objects.select_related('kota_asal', 'kota_tujuan').all()
    if tahun_sbm:
        try:
            tiket_qs = tiket_qs.filter(tahun=int(tahun_sbm))
        except ValueError:
            pass

    if pegawai_id:
        try:
            from master_data.models import Pegawai
            pegawai = Pegawai.objects.get(id=pegawai_id)
            from perjalanan.models import get_eligible_tiket_filter
            tiket_qs = tiket_qs.filter(get_eligible_tiket_filter(pegawai))
        except (ValueError, Pegawai.DoesNotExist):
            pass

    # Order by route fields and -nominal so that deduplication picks the highest rate first
    tiket_qs = tiket_qs.order_by('kota_asal__nama', 'kota_tujuan__nama', 'kelas', '-nominal')

    # Deduplicate route limits
    seen = set()
    dedup_tickets = []
    for t in tiket_qs:
        key = (t.kota_asal_id, t.kota_tujuan_id, t.kelas)
        if key not in seen:
            seen.add(key)
            dedup_tickets.append(t)

    data = []
    for t in dedup_tickets:
        data.append({
            'id': t.id,
            'kota_asal_id': t.kota_asal.id,
            'kota_asal_nama': t.kota_asal.nama,
            'kota_tujuan_id': t.kota_tujuan.id,
            'kota_tujuan_nama': t.kota_tujuan.nama,
            'kelas': t.kelas,
            'kelas_display': t.get_kelas_display(),
            'nominal': float(t.nominal)
        })
    jenis_berkas_tiket_pesawat = list(JenisBerkas.objects.filter(
        kategori_biaya='transportasi_pesawat'
    ).values_list('id', flat=True))
    jenis_berkas_penginapan = list(JenisBerkas.objects.filter(
        kategori_biaya__in=['penginapan', 'penginapan_30', 'penginapan_fb_luar', 'penginapan_fb_dalam']
    ).values_list('id', flat=True))
    return JsonResponse({
        'routes': data,
        'ticket_pesawat_ids': jenis_berkas_tiket_pesawat,
        'penginapan_berkas_ids': jenis_berkas_penginapan
    })

@login_required
def get_standar_biaya_penginapan_ajax(request):
    from master_data.models import StandarBiaya, Pegawai
    from perjalanan.models import get_eligible_sbm_filter
    
    pegawai_id = request.GET.get('pegawai_id')
    provinsi_id = request.GET.get('provinsi_id')
    tahun_sbm = request.GET.get('tahun_sbm')
    perjalanan_id = request.GET.get('perjalanan_id')
    
    try:
        from django.utils import timezone
        from perjalanan.models import PerjalananDinas
        
        if perjalanan_id:
            pd = PerjalananDinas.objects.get(id=perjalanan_id)
            pegawai_id = pd.pegawai_id
            if not tahun_sbm and pd.surat_tugas:
                tahun_sbm = pd.surat_tugas.tahun_sbm
                
        if not all([pegawai_id, provinsi_id]):
            return JsonResponse({'error': 'Missing parameters'}, status=400)
            
        if not tahun_sbm:
            tahun_sbm = timezone.now().year
            
        pegawai = Pegawai.objects.get(id=pegawai_id)
        q_filter = get_eligible_sbm_filter(pegawai)
        sbm = StandarBiaya.objects.filter(
            q_filter,
            provinsi_id=provinsi_id,
            tahun=int(tahun_sbm)
        ).order_by('-plafon_penginapan').first()
        
        plafon = float(sbm.plafon_penginapan) if sbm else 0
        return JsonResponse({'plafon': plafon})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def kalender_perjadin(request):
    import json
    is_admin = request.user.is_staff
    
    if is_admin:
        perjalanan_qs = PerjalananDinas.objects.all().select_related('surat_tugas__tujuan_provinsi', 'pegawai')
    else:
        if not hasattr(request.user, 'pegawai_profile'):
            messages.error(request, "Akun Anda belum terhubung dengan data Pegawai.")
            return redirect('core:dashboard')
        pegawai = request.user.pegawai_profile
        perjalanan_qs = PerjalananDinas.objects.filter(pegawai=pegawai).select_related('surat_tugas__tujuan_provinsi', 'pegawai')
    
    trips = list(perjalanan_qs)
    overlaps = []
    overlapping_ids = set()
    
    from collections import defaultdict
    from datetime import timedelta
    from .models import HarianPerjalanan
    
    # Pre-fetch HarianPerjalanan to avoid N+1 queries
    harian_qs = HarianPerjalanan.objects.filter(perjalanan__in=trips)
    harian_map = {}
    for h in harian_qs:
        harian_map[(h.perjalanan_id, h.tanggal.isoformat())] = h.jenis_harian
    
    # 1. Group travels by pegawai_id
    pegawai_trips = defaultdict(list)
    for t in trips:
        pegawai_trips[t.pegawai_id].append(t)
        
    # 2. Find overlaps per employee grouped by date
    clash_map = {}
    for peg_id, p_trips in pegawai_trips.items():
        date_trips = defaultdict(list)
        for t in p_trips:
            # Find all dates this trip spans
            curr = t.tanggal_berangkat
            if curr and t.tanggal_kembali:
                while curr <= t.tanggal_kembali:
                    date_trips[curr].append(t)
                    curr += timedelta(days=1)
        
        # Filter dates with > 1 trips (conflicts)
        conflict_dates = {}
        for dt, ts in date_trips.items():
            if len(ts) > 1:
                conflict_dates[dt] = ts
                
                # Check resolution: resolved if at most one trip is financed (not 'tidak_dibayai')
                dt_str = dt.isoformat()
                financed_trips = []
                for t in ts:
                    j_harian = harian_map.get((t.id, dt_str), 'luar_kota')
                    if j_harian != 'tidak_dibayai':
                        financed_trips.append(t.id)
                is_resolved = len(financed_trips) <= 1
                clash_map[(peg_id, dt_str)] = {
                    'is_clash': True,
                    'is_resolved': is_resolved,
                    'trips': [t.id for t in ts],
                    'financed_trip_ids': financed_trips
                }
        
        if conflict_dates:
            peg_nama = p_trips[0].pegawai.nama
            overlap_dates_details = []
            for dt in sorted(conflict_dates.keys()):
                trips_on_date = conflict_dates[dt]
                
                trips_details = []
                any_financed = False
                for t in trips_on_date:
                    j_harian = harian_map.get((t.id, dt.isoformat()), 'luar_kota')
                    is_financed = j_harian != 'tidak_dibayai'
                    
                    overlapping_ids.add(t.id)
                    
                    trips_details.append({
                        'id': t.id,
                        'nomor_spd': t.nomor_spd or 'Draft SPD',
                        'tujuan': t.tujuan_provinsi.nama,
                        'perihal': t.surat_tugas.perihal,
                        'maksud_perjalanan': t.maksud_perjalanan,
                        'is_financed': is_financed
                    })
                    if is_financed:
                        any_financed = True
                        
                overlap_dates_details.append({
                    'tanggal_iso': dt.isoformat(),
                    'tanggal_str': dt.strftime('%d %b %Y'),
                    'trips': trips_details,
                    'none_financed': not any_financed
                })
            
            overlaps.append({
                'pegawai_id': peg_id,
                'pegawai_nama': peg_nama,
                'conflict_dates': overlap_dates_details
            })
                
    serialized_trips = []
    for t in trips:
        dates_info = {}
        curr = t.tanggal_berangkat
        if curr and t.tanggal_kembali:
            while curr <= t.tanggal_kembali:
                dt_str = curr.isoformat()
                j_harian = harian_map.get((t.id, dt_str), 'luar_kota')
                clash_info = clash_map.get((t.pegawai_id, dt_str))
                
                is_overlap = False
                is_unresolved_clash = False
                if clash_info:
                    if not clash_info['is_resolved']:
                        is_overlap = True
                        is_unresolved_clash = True
                    else:
                        if j_harian == 'tidak_dibayai':
                            is_overlap = True
                
                dates_info[dt_str] = {
                    'jenis_harian': j_harian,
                    'is_overlap': is_overlap,
                    'is_unresolved_clash': is_unresolved_clash
                }
                curr += timedelta(days=1)

        serialized_trips.append({
            'id': t.id,
            'nomor_spd': t.nomor_spd or 'Draft SPD',
            'nomor_surat': t.surat_tugas.nomor_surat,
            'perihal': t.surat_tugas.perihal,
            'pegawai_nama': t.pegawai.nama,
            'pegawai_id': str(t.pegawai.id),
            'tanggal_berangkat': t.tanggal_berangkat.isoformat() if t.tanggal_berangkat else None,
            'tanggal_kembali': t.tanggal_kembali.isoformat() if t.tanggal_kembali else None,
            'tujuan': t.tujuan_provinsi.nama,
            'status': t.status,
            'status_display': t.get_status_display(),
            'has_overlap': t.id in overlapping_ids,
            'dates_info': dates_info
        })
        
    context = {
        'trips_json': json.dumps(serialized_trips),
        'overlaps': overlaps,
        'has_overlaps': len(overlaps) > 0,
        'is_admin': is_admin,
    }
    return render(request, 'perjalanan/kalender_perjadin.html', context)


@login_required
@staff_member_required
def resolusi_konflik(request):
    import datetime
    from .models import HarianPerjalanan
    
    if request.method == 'POST':
        pegawai_id = request.POST.get('pegawai_id')
        
        # We find all inputs chosen_<date_iso>
        for key, value in request.POST.items():
            if key.startswith('chosen_'):
                date_str = key.replace('chosen_', '')
                try:
                    tanggal = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    continue
                chosen_trip_id = value
                
                # Get all HarianPerjalanan for this pegawai and date
                harian_days = HarianPerjalanan.objects.filter(
                    perjalanan__pegawai_id=pegawai_id,
                    tanggal=tanggal
                )
                
                for h in harian_days:
                    if str(h.perjalanan_id) == chosen_trip_id:
                        # Restore to luar_kota if it was set to tidak_dibayai
                        if h.jenis_harian == 'tidak_dibayai':
                            h.jenis_harian = 'luar_kota'
                            h.save()
                    else:
                        # Set to tidak_dibayai
                        if h.jenis_harian != 'tidak_dibayai':
                            h.jenis_harian = 'tidak_dibayai'
                            h.save()
                            
        messages.success(request, "Resolusi bentrok jadwal berhasil disimpan.")
    return redirect('perjalanan:kalender_perjadin')


import os
import openpyxl
from django.conf import settings
from django.http import HttpResponse
from master_data.models import PejabatPenandatangan

@login_required
def download_spd_excel(request, perjadin_id):
    perjadin = get_object_or_404(PerjalananDinas, id=perjadin_id)
    surat_tugas = perjadin.surat_tugas
    pegawai = perjadin.pegawai

    months = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
        7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }

    def format_date_indo(date_obj):
        if not date_obj:
            return ""
        return f"{date_obj.day} {months[date_obj.month]} {date_obj.year}"

    ppk = PejabatPenandatangan.objects.filter(jabatan='PPK').first()
    nama_ppk = ppk.nama if ppk else "-"
    nip_ppk = ppk.nip if ppk else "-"
    
    # Ambil durasi hari
    durasi = perjadin.durasi_hari
    if durasi is None:
        durasi = "-"
        
    # Ambil tempat tujuan
    tempat_tujuan = surat_tugas.tujuan_provinsi.nama if surat_tugas.tujuan_provinsi else surat_tugas.tempat_tujuan

    mapping = {
        '{{FULL_NO_SPD}}': perjadin.nomor_spd or "-",
        '{{NO_SPD}}': perjadin.nomor_spd.split('/')[0] if perjadin.nomor_spd else "-",
        '{{NAMA_PEGAWAI}}': pegawai.nama,
        '{{NIP_PEGAWAI}}': pegawai.nip if pegawai.nip else "-",
        '{{GOLONGAN_PEGAWAI}}': pegawai.get_golongan_display(),
        '{{JABATAN_PEGAWAI}}': pegawai.jabatan,
        '{{MAKSUD_PERJALANAN_DINAS}}': surat_tugas.maksud_perjalanan or surat_tugas.perihal,
        '{{JENIS_TRANSPORTASI}}': surat_tugas.get_jenis_transportasi_display(),
        '{{TEMPAT_BERANGKAT}}': surat_tugas.tempat_berangkat or "-",
        '{{TEMPAT_TUJUAN}}': tempat_tujuan or "-",
        '{{LAMA_HARI_PERJADIN}}': f"{durasi} Hari",
        '{{TANGGAL_BERANGKAT}}': format_date_indo(surat_tugas.tanggal_berangkat),
        '{{TANGGAL_KEMBALI}}': format_date_indo(surat_tugas.tanggal_kembali),
        '{{TANGGAL_SPD_DIBUAT}}': format_date_indo(perjadin.created_at.date() if perjadin.created_at else None),
        '{{NAMA_PPK}}': nama_ppk,
        '{{NIP_PPK}}': nip_ppk,
        '{{SUMBER_ANGGARAN}}': f"{surat_tugas.anggaran.kode_dipa} - {surat_tugas.anggaran.nama_kegiatan}" if surat_tugas.anggaran else "-",
    }

    template_path = os.path.join(settings.BASE_DIR, 'static', 'template_documents', 'template_SPD.xlsx')
    
    import zipfile
    import io

    # Gunakan in-memory buffer untuk menyimpan hasil modifikasi zip
    output_buffer = io.BytesIO()
    
    with zipfile.ZipFile(template_path, 'r') as zin:
        with zipfile.ZipFile(output_buffer, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == 'xl/sharedStrings.xml' or item.filename.startswith('xl/worksheets/'):
                    text = content.decode('utf-8')
                    for key, val in mapping.items():
                        text = text.replace(key, str(val))
                    content = text.encode('utf-8')
                zout.writestr(item, content)

    response = HttpResponse(
        output_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="SPD_{pegawai.nama}.xlsx"'
    return response

@login_required
def download_rincian_excel(request, perjadin_id):
    perjadin = get_object_or_404(PerjalananDinas, id=perjadin_id)
    surat_tugas = perjadin.surat_tugas

    months = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
        7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }

    def format_date_indo(date_obj):
        if not date_obj:
            return ""
        return f"{date_obj.day} {months[date_obj.month]} {date_obj.year}"

    ppk = PejabatPenandatangan.objects.filter(jabatan='PPK').first()
    nama_ppk = ppk.nama if ppk else "-"
    nip_ppk = ppk.nip if ppk else "-"

    sekretaris = PejabatPenandatangan.objects.filter(jabatan='SEKRETARIS').first()
    nama_sekretaris = sekretaris.nama if sekretaris else "-"
    nip_sekretaris = sekretaris.nip if sekretaris else "-"

    bendahara = PejabatPenandatangan.objects.filter(jabatan='BENDAHARA').first()
    nama_bendahara = bendahara.nama if bendahara else "-"
    nip_bendahara = bendahara.nip if bendahara else "-"
    
    # Hitung total biaya
    try:
        biaya = perjadin.biaya
        breakdown = biaya.calculate_breakdown()
        total = breakdown.get('total_dibayarkan', 0)
    except Exception:
        biaya = None
        breakdown = {}
        total = 0
        
    formatted_total = '{:,.0f}'.format(total).replace(',', '.')

    # Tambahan TAHUN_BULAN_SEKARANG
    from datetime import date
    today = date.today()
    tahun_bulan_sekarang = f"{months[today.month]} {today.year}"

    mapping = {
        '{{NOMOR_SPD}}': perjadin.nomor_spd or "-",
        '{{TANGGAL_SURAT_SPD}}': format_date_indo(perjadin.created_at.date() if perjadin.created_at else None),
        '{{NAMA_PPK}}': nama_ppk,
        '{{NIP_PPK}}': nip_ppk,
        '{{NAMA_SEKRETARIS}}': nama_sekretaris,
        '{{NIP_SEKRETARIS}}': nip_sekretaris,
        '{{NAMA_BENDAHARA}}': nama_bendahara,
        '{{NIP_BENDAHARA}}': nip_bendahara,
        '{{TOTAL_BIAYA}}': formatted_total,
        '{{TOTAL_DIBAYARKAN}}': formatted_total,
        '{{TAHUN_BULAN_SEKARANG}}': tahun_bulan_sekarang,
        '{{BULAN_TAHUN_SEKARANG}}': tahun_bulan_sekarang,
    }

    template_path = os.path.join(settings.BASE_DIR, 'static', 'template_documents', 'rincian_SPD.xlsx')
    
    import openpyxl
    from copy import copy
    import io
    import zipfile
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 1. Replace variables in all cells
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for key, val in mapping.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, str(val))
                        
    # Kumpulkan item rincian
    items_to_insert = []
    cats = breakdown.get('breakdown_categories', {})
    for cat_key, cat_data in cats.items():
        for item in cat_data.get('items', []):
            if item.get('total', 0) > 0:
                items_to_insert.append({
                    'uraian': item.get('keterangan', ''),
                    'jumlah': item.get('total', 0)
                })

    start_row = 12
    # Fill data into fixed rows
    for i, item in enumerate(items_to_insert):
        row_idx = start_row + i
        if row_idx > 28:  # Batas maksimal row rincian yang ada di template
            break
            
        formatted_jumlah = 'Rp {:,.0f}'.format(item['jumlah']).replace(',', '.')
        
        ws.cell(row=row_idx, column=2, value=str(i+1))
        ws.cell(row=row_idx, column=3, value=item['uraian'])
        ws.cell(row=row_idx, column=7, value=formatted_jumlah)

    openpyxl_buf = io.BytesIO()
    wb.save(openpyxl_buf)
    openpyxl_buf.seek(0)
    
    # 2. Re-inject drawing tag to preserve the image
    final_buf = io.BytesIO()
    with zipfile.ZipFile(openpyxl_buf, 'r') as zin:
        with zipfile.ZipFile(final_buf, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == 'xl/worksheets/sheet1.xml':
                    text = content.decode('utf-8')
                    if '<drawing' not in text:
                        text = text.replace('</worksheet>', '<drawing r:id="rId1"/></worksheet>')
                    content = text.encode('utf-8')
                zout.writestr(item, content)

    response = HttpResponse(
        final_buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Rincian_Biaya_{perjadin.pegawai.nama}.xlsx"'
    return response

@login_required
def download_kwitansi_excel(request, perjadin_id):
    perjadin = get_object_or_404(PerjalananDinas, id=perjadin_id)
    surat_tugas = perjadin.surat_tugas
    pegawai = perjadin.pegawai

    months = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
        7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }

    def format_date_indo(date_obj):
        if not date_obj:
            return ""
        return f"{date_obj.day} {months[date_obj.month]} {date_obj.year}"

    from datetime import date
    today = date.today()
    tanggal_kwitansi = f"{today.day} {months[today.month]} {today.year}"

    ppk = PejabatPenandatangan.objects.filter(jabatan='PPK').first()
    nama_ppk = ppk.nama if ppk else "-"
    nip_ppk = ppk.nip if ppk else "-"

    try:
        biaya = perjadin.biaya
        breakdown = biaya.calculate_breakdown()
        total = breakdown.get('total_dibayarkan', 0)
    except Exception:
        total = 0
        
    formatted_total = '{:,.0f}'.format(total).replace(',', '.')

    tempat_tujuan = surat_tugas.tujuan_provinsi.nama if surat_tugas.tujuan_provinsi else surat_tugas.tempat_tujuan

    mapping = {
        '{{NO_FULL_SPD}}': perjadin.nomor_spd or "-",
        '{{NAMA_PEGAWAI}}': pegawai.nama,
        '{{NIP_PEGAWAI}}': pegawai.nip if pegawai.nip else "-",
        '{{NAMA_PPK}}': nama_ppk,
        '{{NIP_PPK}}': nip_ppk,
        '{{TANGGAL_KWITANSI}}': tanggal_kwitansi,
        '{{TANGGAL_SURAT_SPD}}': format_date_indo(perjadin.created_at.date() if perjadin.created_at else None),
        '{{TEMPAT_BERANGKAT}}': surat_tugas.tempat_berangkat or "-",
        '{{TEMPAT_TUJUAN}}': tempat_tujuan or "-",
        '{{TOTAL_DIBAYARKAN}}': formatted_total,
    }

    template_path = os.path.join(settings.BASE_DIR, 'static', 'template_documents', 'kwitansi_SPD.xlsx')
    
    import zipfile
    import io

    output_buffer = io.BytesIO()
    
    with zipfile.ZipFile(template_path, 'r') as zin:
        with zipfile.ZipFile(output_buffer, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == 'xl/sharedStrings.xml' or item.filename.startswith('xl/worksheets/'):
                    text = content.decode('utf-8')
                    for key, val in mapping.items():
                        text = text.replace(key, str(val))
                    content = text.encode('utf-8')
                zout.writestr(item, content)

    response = HttpResponse(
        output_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Kwitansi_{pegawai.nama}.xlsx"'
    return response
